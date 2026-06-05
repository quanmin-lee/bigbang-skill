"""Core module for Excel summary CLI tool."""

import os
import statistics
import openpyxl


_EMPTY_SUMMARY = {"sum": 0, "avg": 0, "max": 0, "min": 0, "count": 0}


def _get_column_index(header, column: str) -> int:
    """Find the column index for the given column name in the header row.

    Args:
        header: Tuple of header cell values.
        column: Column name to find.

    Returns:
        Zero-based column index.

    Raises:
        ValueError: If the column is not found.
    """
    if column in header:
        return header.index(column)
    raise ValueError(f"Column '{column}' not found in spreadsheet")


def read_excel(path: str, column: str) -> list[float]:
    """Read an Excel file and extract numeric values from the specified column.

    Args:
        path: Path to the .xlsx file.
        column: Column name to extract numeric values from.

    Returns:
        List of float values found in the column.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the column is not found in the worksheet.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find header row and locate column index
    header = next(ws.iter_rows(values_only=True), None)
    if header is None:
        raise ValueError(f"Column '{column}' not found in spreadsheet")
    col_idx = _get_column_index(header, column)

    # Extract numeric values from the column (skip header)
    values: list[float] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col_idx < len(row):
            cell = row[col_idx]
            if isinstance(cell, (int, float)):
                values.append(float(cell))

    return values


def summarize_column(values: list[float]) -> dict[str, float | int]:
    """Compute summary statistics for a list of numeric values.

    Args:
        values: List of float values to summarize.

    Returns:
        Dictionary with keys: sum, avg, max, min, count (unique count).
    """
    if not values:
        return _EMPTY_SUMMARY.copy()

    total = sum(values)
    avg = statistics.mean(values)
    max_val = max(values)
    min_val = min(values)
    unique_count = len(set(values))

    return {
        "sum": total,
        "avg": avg,
        "max": max_val,
        "min": min_val,
        "count": unique_count,
    }
